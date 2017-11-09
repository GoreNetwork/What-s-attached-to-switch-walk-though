#from excel_work import*
from common_functions import *
from pull import *
#from mac_and_arp_work import *
from napalm import get_network_driver	
from getpass import getpass
from pprint import pprint
from name_work import *
import openpyxl

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

def write_excel_data(row,column,value,sheet):
	tmp = str(get_column_letter(column))+str(row)
	sheet[tmp] = value
	return column+1


username = input("Username: ")
password = getpass()

def normalize_mac (mac):
	mac = mac.strip(" ")
	mac = mac.replace('.','')
	mac = mac.upper()
	t = iter(mac)
	mac = ':'.join(a+b for a,b in zip(t, t))
	return mac
	

def pull_mac_table (netconnect):
	mac_table_list = []
	mac_table = net_connect.send_command_expect('sh mac address-table')
	for line in mac_table.split('\n'):
		mac_int = {}
		
		line = line.lstrip(" ")
		line = line.rstrip(" ")
		if len(get_mac (line)) == 0:
			continue
		mac = normalize_mac (get_mac (line)[0])
		if mac =='FF:FF:FF:FF:FF:FF':
			continue
		mac_int['mac']= mac
		#print (line.split(" ")[-1])
		interface = normalize_interface_names(line.split(" ")[-1])
		mac_int['interface'] = interface
		mac_table_list.append(mac_int)
	return mac_table_list

	
def ouicorrect(list):
	templist = []
	for oui in list:
		templist.append(oui[0:7])
	return (templist)	
	
def ouicorrect(list):
	templist = []
	for oui in list:
		templist.append(normalize_mac (oui[0:7]))
	return (templist)	
	
def check_ouis(folder_name):
	os.chdir(folder_name)	
	files = os.listdir()
	#print (files)
	OUIs = {}	
	for file in files:
		q =open(file).readlines()
		fixed_oui = ouicorrect (q)
		for each_oui in fixed_oui:
			OUIs [each_oui]= file
	os.chdir(os.pardir)
	return OUIs

	
def pull_oui_type(mac_address,OUIs):
	mac_oui = mac_address[0:8]
	if mac_oui in OUIs:
		#print ("ITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKEDITWORKED")
		return OUIs[mac_oui]
	else:
		#print (mac_oui)
		return "Unknown"
	
	
driver = get_network_driver('ios')

to_check = read_doc ('pull_these.csv')
cdp_file_name= "temp_cdp_info.txt"




##device = driver("10.9.106.238", username,password)
#device.open() 
#pprint(device.get_mac_address_table())
interfaces =[]

wb = openpyxl.Workbook()


folder_name = "OUIs"
OUIs = check_ouis(folder_name)

for device_and_arp in to_check:
	print (device_and_arp)
#	try:
	interfaces = []
	switch_to_check = device_and_arp.split(',')[0]
	arp_device = device_and_arp.split(',')[1]
	arp_device = arp_device.rstrip("\n")
	print ("working on "+switch_to_check )
	driver = get_network_driver('ios')
	device = driver(arp_device, username,password)
	device.open() 

	arp_table = device.get_arp_table()
	net_connect = make_connection (switch_to_check,username,password)
	mac_table = pull_mac_table (net_connect)
	#pprint (mac_table )
	#pprint(arp_table)
	for mac_entry in mac_table:
		for arp_entry in arp_table:
			if mac_entry['mac'] == arp_entry['mac']:
				tmp = {}
				tmp['mac'] = mac_entry['mac']
				tmp['interface'] = mac_entry['interface']
				tmp['ip'] = arp_entry['ip']
				tmp['type'] = pull_oui_type(tmp['mac'],OUIs)
				try:
					tmp['hostname'] = socket.gethostbyaddr(tmp['ip'])[0]
					#print (tmp['hostname'])
				except:
					tmp['hostname'] = "Unknown"
				interfaces.append(tmp)
	#pprint (interfaces)
	sheet = wb.create_sheet(switch_to_check)
	row = 1
	for interface in interfaces:
		column = 1
		if interface['interface'] == "Switch":
			continue
		column = write_excel_data(row,column,interface['interface'],sheet)
		column = write_excel_data(row,column,interface['ip'],sheet)
		column = write_excel_data(row,column,interface['type'],sheet)
		column = write_excel_data(row,column,interface['hostname'],sheet)
		column = write_excel_data(row,column,interface['mac'],sheet)
		row = row+1
	row = row+1
	file_name = switch_to_check+ " show cdp"
	pull_cdp_output(switch_to_check,username,password,file_name)
	#cdp_info = cdp_info.split('\n')
	for each in parse_cdp_out(file_name):
		column = 1
		column = write_excel_data(row,column,each['remote_id'],sheet)
		column = write_excel_data(row,column,each['remote_ip'],sheet)
		column = write_excel_data(row,column,each['local_int'],sheet)
		column = write_excel_data(row,column,each['remote_int'],sheet)
		column = write_excel_data(row,column,each['platform'],sheet)
		row = row+1
#	except:
#		print(switch_to_check+ " Didn't work"+switch_to_check+ " Didn't work"+switch_to_check+ " Didn't work"+switch_to_check+ " Didn't work")
	
	
wb.save('output.xlsx')
	
			
	
	
	
